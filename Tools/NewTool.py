import PySimpleGUI as sg
import json
import openpyxl

with open(r'N:\Images\Alisa\AgeEstimate\Jsons\000E1712B4624ED09EC0284A74953D8E_ID\Data.json') as json_file:
    # Load the JSON data
    data = json.load(json_file)


    # Define a recursive function to extract the values and keys
    def extract_values(data, result, key=None):
        # If the data is a dictionary, extract the values and keys from its keys and values
        if isinstance(data, dict):
            for k, value in data.items():
                extract_values(value, result, k)
        # If the data is a list, extract the values and keys from its elements
        elif isinstance(data, list):
            for item in data:
                extract_values(item, result, key)
        # If the data is a leaf node, add the value and key to the result dictionary
        else:
            result[key] = data


    # Initialize an empty dictionary to store the extracted values and keys
    result = {}

    # Extract the values and keys from the JSON data
    extract_values(data, result)

    # Create the GUI
    layout = [
        [sg.Text('Select the values to extract:')],
        [sg.Listbox(values=list(result.keys()), size=(30, 10), key='selected_values', select_mode=sg.SELECT_MODE_MULTIPLE)],
        [sg.Button('Extract'), sg.Button('Cancel')]
    ]
    window = sg.Window('JSON Extractor', layout)

    # Read the values selected by the user
    while True:
        event, values = window.read()
        if event in (None, 'Cancel'):
            break
        elif event == 'Extract':
            # Create the Excel file
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Extracted Data'
            search_list = values['selected_values']
            print(search_list)
            # for field_name in values['selected_values']:
            #     print(field_name)
            for search_key in values['selected_values']:
                print(search_key)
                # print(values['selected_values'])
                print(result[search_key])
                # Write the selected values and field names to the Excel file
                # print(values['selected_values'])
                # for i, key in enumerate(values['selected_values'], start=1):
                #     # print(values[i - 1])
                #     # print(key)
                #     # print(values[i - 1])
                #     sheet.cell(row=i, column=1).value = str(names[i - 1])
                #
                #     # sheet.cell(row=i, column=2).value = str(values[i - 1])
                #
                # # Save the Excel file
                # wb.save('extracted_data.xlsx')

            # Show a success message
            sg.popup('Data extracted successfully!')

            # Close the window
            break

# Close the window
window.close()

# import json
#
# # Open the JSON file
# with open(r'N:\Images\Alisa\AgeEstimate\Jsons\000E1712B4624ED09EC0284A74953D8E_ID\Data.json') as json_file:
#     # Load the JSON data
#     data = json.load(json_file)
#
#
#     # Define a recursive function to extract the values and keys
#     def extract_values(data, result, key=None):
#         # If the data is a dictionary, extract the values and keys from its keys and values
#         if isinstance(data, dict):
#             for k, value in data.items():
#                 extract_values(value, result, k)
#         # If the data is a list, extract the values and keys from its elements
#         elif isinstance(data, list):
#             for item in data:
#                 extract_values(item, result, key)
#         # If the data is a leaf node, add the value and key to the result dictionary
#         else:
#             result[key] = data
#
#
#     # Initialize an empty dictionary to store the extracted values and keys
#     result = {}
#
#     # Extract the values and keys from the JSON data
#     extract_values(data, result)
#
#     # Print the result dictionary
#     print(result)
